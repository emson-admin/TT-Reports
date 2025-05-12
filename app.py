import streamlit as st
import pandas as pd
import datetime
import gspread
import re
import io
from google.oauth2.service_account import Credentials
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from google.cloud import storage
import uuid
import requests
import altair as alt # Moved import altair to top
import os

#page layout
st.set_page_config(layout="wide")

# --- Helper Functions ---
def extract_account(campaign_name):
    if pd.isna(campaign_name):
        return "Other Accounts"
    
    campaign_name_lower = str(campaign_name).lower() # Convert to string and lowercase for case-insensitive search

    if "granitestone" in campaign_name_lower:
        return "Granitestone"
    elif "bell and howell" in campaign_name_lower: # Search for "bell and howell"
        return "Bell+Howell" # Return "Bell+Howell"
    
    return "Other Accounts" # Fallback for campaigns not matching the keywords

def upload_excel_to_gcs(excel_bytes, bucket_name, credentials_dict, original_filename="report.xlsx"):
    """Uploads an Excel file (in bytes) to Google Cloud Storage."""
    try:
        buffer = io.BytesIO(excel_bytes)
        buffer.seek(0)

        # Create a unique filename, perhaps incorporating the original name
        base, ext = os.path.splitext(original_filename)
        filename = f"reports/{base}_{uuid.uuid4().hex}{ext}"

        # Init GCS client
        client = storage.Client.from_service_account_info(credentials_dict)
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(filename)
        blob.upload_from_file(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Construct the public URL manually for uniform bucket-level access
        public_url = f"https://storage.googleapis.com/{bucket.name}/{blob.name}"
        return public_url
    except Exception as e:
        st.error(f"Error uploading Excel to GCS: {e}")
        return None

def upload_chart_to_gcs(chart, bucket_name, credentials_dict):
    """Uploads an Altair chart image to Google Cloud Storage."""
    try:
        # Convert chart to PNG in memory
        buffer = io.BytesIO()
        # Ensure chart object is valid before saving
        if chart is None:
            st.error("Chart object is None, cannot upload.")
            return None
        chart.save(buffer, format='png', scale_factor=2.0)
        buffer.seek(0)

        # Create a unique filename
        filename = f"charts/chart_{uuid.uuid4().hex}.png"

        # Init GCS client
        client = storage.Client.from_service_account_info(credentials_dict)
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(filename)
        blob.upload_from_file(buffer, content_type='image/png')

        # Make it publicly accessible - REMOVED due to Uniform Bucket-Level Access
        # blob.make_public() 
        
        # Construct the public URL manually for uniform bucket-level access
        # This assumes the bucket is configured for public read access in GCP IAM.
        public_url = f"https://storage.googleapis.com/{bucket.name}/{blob.name}"
        return public_url
    except Exception as e:
        st.error(f"Error uploading chart to GCS: {e}")
        return None

def generate_specific_metric_chart(chart_df, metric_col_name, display_metric_name, date_col='report_date', campaign_col='campaign_name'):
    """Generates an Altair chart for a single metric using pre-aggregated data."""
    if metric_col_name not in chart_df.columns:
        st.warning(f"Metric column '{metric_col_name}' not found in chart data for '{display_metric_name}'.")
        return None
    
    chart = alt.Chart(chart_df).mark_line(point=True).encode(
        x=alt.X(f'{date_col}:T', title='Date'),
        y=alt.Y(f'{metric_col_name}:Q', title=display_metric_name.replace("_", " ").title(), scale=alt.Scale(zero=True)),
        color=alt.Color(f'{campaign_col}:N', legend=alt.Legend(title="Campaign")),
        tooltip=[
            alt.Tooltip(f'{date_col}:T', title="Date"),
            alt.Tooltip(f'{campaign_col}:N', title="Campaign"),
            alt.Tooltip(f'{metric_col_name}:Q', title=display_metric_name.replace("_", " ").title(), format=",.2f")
        ]
    ).properties(
        width=600,
        height=300,
        title=f"{display_metric_name.replace('_', ' ').title()} Over Time"
    ).interactive()
    return chart

# --- Main Application Logic ---
def run_main_app():
    # Connect to Google Sheets
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = Credentials.from_service_account_info(st.secrets["gcp"], scopes=scope)
    gc = gspread.authorize(credentials)
    sh = gc.open('TikTok GMVMAX Ad Reports').worksheet('Data')

    st.title('📈 Weekly Ad Report Uploader')

    # If flag is set, clear uploader state by resetting its key
    if st.session_state.get("clear_uploader"):
        st.session_state.pop("uploader", None)
        st.session_state["clear_uploader"] = False

    # File uploader
    if st.session_state.get("is_admin"):
        st.markdown("## 🔼 Upload Daily Reports")
        uploaded_files = st.file_uploader(
            "📂 Upload one or more daily ad reports (.xlsx)",
            type="xlsx",
            accept_multiple_files=True,
            key="uploader"
        )

        # Define all_data globally so it's always initialized
        all_data = []
        if uploaded_files:

            for file in uploaded_files:
                try:
                    df = pd.read_excel(file)

                    # Clean & normalize
                    df.columns = df.columns.astype(str).str.strip().str.lower().str.replace(" ", "_")
                    
                    # Parse date from filename
                    date_match = re.search(r"\d{4}-\d{2}-\d{2}", file.name)
                    if date_match:
                        report_date = pd.to_datetime(date_match.group())
                        df['report_date'] = report_date
                    else:
                        st.warning(f"⚠️ Could not find date in: {file.name}")
                        continue

                    # Remove zero-spend rows
                    if 'cost' in df.columns:
                        df = df[df['cost'].fillna(0) > 0]

                    # Convert campaign_id to string
                    if 'campaign_id' in df.columns:
                        df['campaign_id'] = df['campaign_id'].apply(lambda x: f"'{x}")

                    df['upload_date'] = datetime.datetime.now().strftime('%Y-%m-%d')

                    all_data.append(df)

                except Exception as e:
                    st.error(f"❌ Error processing {file.name}: {e}")

            if all_data and st.button("✅ Upload All to Google Sheets"):
                final_df = pd.DataFrame() # Initialize final_df as an empty DataFrame

                # Combine uploaded data into one DataFrame
                combined_df = pd.concat(all_data, ignore_index=True)

                # Add "account_name" column using the helper function (normalized name)
                if 'campaign_name' in combined_df.columns:
                    combined_df['account_name'] = combined_df['campaign_name'].apply(extract_account)
                else:
                    # If campaign_name is somehow missing in uploaded data, fill with default
                    combined_df['account_name'] = "Other Accounts" 

                # Load existing data from sheet
                existing = pd.DataFrame(sh.get_all_records())
                existing.columns = pd.Index([str(col).strip().lower().replace(" ", "_") for col in existing.columns])

                # Ensure combined_df['report_date'] exists and is datetime
                combined_df['report_date'] = pd.to_datetime(combined_df['report_date'])
                if not existing.empty and 'report_date' in existing.columns:
                    existing['report_date'] = pd.to_datetime(existing['report_date'])

                    # Normalize campaign_id columns for reliable joins
                    if 'campaign_id' in combined_df.columns and 'campaign_id' in existing.columns:
                        existing['campaign_id'] = existing['campaign_id'].astype(str).str.strip()
                        combined_df['campaign_id'] = combined_df['campaign_id'].astype(str).str.strip()

                        # Find rows that would be duplicates (same campaign_id + report_date)
                        duplicate_rows = pd.merge(
                            combined_df,
                            existing[['campaign_id', 'report_date']],
                            on=['campaign_id', 'report_date'],
                            how='inner'
                        )

                        # If duplicates exist, let user choose whether to skip or overwrite them
                        if not duplicate_rows.empty:
                            with st.expander("⚠️ Duplicate Rows Detected"):
                                st.dataframe(duplicate_rows)

                            # Ask user whether to overwrite or skip duplicate rows
                            overwrite = st.radio(
                                "⚠️ Duplicate rows detected based on 'campaign_id' and 'report_date'. What would you like to do?",
                                ("Skip duplicates", "Overwrite duplicates")
                            )

                            if overwrite == "Overwrite duplicates":
                                # Remove duplicates from the existing sheet data
                                existing = existing.merge(
                                    duplicate_rows[['campaign_id', 'report_date']],
                                    on=['campaign_id', 'report_date'],
                                    how='left',
                                    indicator=True
                                ).query('_merge == "left_only"').drop(columns=['_merge'])

                                # Final data is: cleaned existing rows + all uploaded rows (including overwrites)
                                final_df = pd.concat([existing, combined_df], ignore_index=True)
                                st.info(f"✅ Overwrote {len(duplicate_rows)} existing rows.")
                            else:
                                # Skip: remove duplicate rows from the upload
                                final_df = pd.merge(
                                    combined_df,
                                    duplicate_rows[['campaign_id', 'report_date']],
                                    on=['campaign_id', 'report_date'],
                                    how='left',
                                    indicator=True
                                ).query('_merge == "left_only"').drop(columns(['_merge']))
                                st.info(f"✅ Skipped {len(duplicate_rows)} duplicate rows.")
                        else:
                            # No duplicates: just combine everything
                            final_df = pd.concat([existing, combined_df], ignore_index=True)
                    else:
                        # Fallback if campaign_id missing: dedupe just by date
                        final_df = combined_df[~combined_df['report_date'].isin(existing['report_date'])]
                        final_df = pd.concat([existing, final_df], ignore_index=True)
                        st.info(f"⚠️ No 'campaign_id' column found. Deduped only by date.")
                else:
                    # No existing data at all — treat this as the first upload
                    final_df = combined_df.copy()
                    st.info("ℹ️ No existing data found. Treating this as a fresh upload.")

                # At this point, final_df should have been assigned by one of the conditional paths.
                # If it's still the initial empty DataFrame and it shouldn't be, that's a deeper logic issue.
                # But this initialization prevents the UnboundLocalError.

                # Get current sheet row count to determine where to append
                existing_records = sh.get_all_values()
                is_first_upload = len(existing_records) == 0

                # Convert any datetime columns to string format for JSON compatibility with gspread
                for col in final_df.select_dtypes(include=['datetime', 'datetime64[ns]']):
                    final_df[col] = final_df[col].dt.strftime('%Y-%m-%d')


                # Build the upload payload
                rows_to_upload = (
                    [final_df.columns.tolist()] + final_df.astype(str).values.tolist()
                    if is_first_upload else final_df.astype(str).values.tolist()
                )


                # Determine next row to write to
                next_row = len(existing_records) + 1

                # Upload to Google Sheet
                sh.update(f'A{next_row}', rows_to_upload)

                # Show success and rerun app to reflect updated data
                st.success(f'🎉 Uploaded {len(uploaded_files)} file(s) successfully!')
                st.toast("Upload complete!", icon="✅")
                # Clear uploader input by resetting session state
                st.session_state["clear_uploader"] = True
                st.rerun()
            
    @st.cache_data(ttl=0)
    def load_data():
        df = pd.DataFrame(sh.get_all_records())
        if not df.empty:
            df.columns = pd.Index([str(col).strip().lower().replace(" ", "_") for col in df.columns])
            
            # Define columns that should be numeric
            # Add any other columns that are expected to be numeric from your sheet
            cols_to_convert_to_numeric = [
                'cost', 'gross_revenue', 'orders_(sku)', 'roi', 'cost_per_order',
                'impressions', 'clicks', 'ctr', 'cpc', 'cpm', 'net_cost' # Add other potential numeric metrics
            ]
            
            for col in cols_to_convert_to_numeric:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Ensure report_date is datetime (moved here for early conversion)
            if 'report_date' in df.columns:
                 df['report_date'] = pd.to_datetime(df['report_date'], errors='coerce')
            else:
                # If 'report_date' is critical and missing, you might want to handle it,
                # e.g., by creating an empty column or raising an error earlier.
                # For now, the downstream check will catch it.
                pass

        return df

    # Load data from Google Sheets
    data = load_data()

    # Make sure report_date is datetime (this check can be more robust now or rely on load_data)
    if 'report_date' not in data.columns or data['report_date'].isnull().all(): # Check if column exists and is not all NaT
        st.error("❌ The 'report_date' column is missing or contains no valid dates. Double-check the upload formatting.")
        st.stop()
    # data['report_date'] = pd.to_datetime(data['report_date']) # This line is now handled in load_data

    # --- Account Column Handling (uses 'Account Name' from sheet, normalized to 'account_name') ---
    # Ensure 'account_name' column exists for filtering, especially for backward compatibility or empty sheet
    if data.empty:
        # If data is empty, create placeholder columns to prevent downstream errors
        # 'report_date' is already handled by an st.stop() if missing/empty
        data['campaign_name'] = pd.Series(dtype='object') 
        data['account_name'] = pd.Series(dtype='object') 
    elif 'account_name' not in data.columns:
        # Data is not empty, but 'account_name' column is missing (e.g. old data)
        st.sidebar.info("Deriving 'account_name' for older data or missing column.")
        if 'campaign_name' in data.columns:
            data['account_name'] = data['campaign_name'].apply(extract_account)
        else:
            # If 'campaign_name' is also missing, fill 'account_name' with default
            data['account_name'] = pd.Series(["Other Accounts"] * len(data), index=data.index)


    with st.sidebar:
        st.markdown("### 🔍 Filter Options")
        
        # Date Range Filter - ensure data is not empty before trying to access min/max
        if not data.empty and 'report_date' in data.columns:
            min_date_val = data['report_date'].min().date()
            max_date_val = data['report_date'].max().date()
        else: # Default if data is empty or report_date is missing (though latter should be caught by st.stop)
            min_date_val = datetime.date.today() - datetime.timedelta(days=7)
            max_date_val = datetime.date.today()
        
        date_range = st.date_input("Report Date Range", value=(min_date_val, max_date_val))
        
        # Account Filter - uses 'account_name'
        if not data.empty and 'account_name' in data.columns:
            # Ensure all items are strings before sorting
            unique_accounts = [str(acc) for acc in data['account_name'].unique().tolist()]
            account_options = ["All Accounts"] + sorted(unique_accounts)
            selected_accounts = st.multiselect("Account(s)", account_options, default=[]) # Default to empty list
        else:
            selected_accounts = [] # Default to empty list

        # --- Derive Campaign Options based on Selected Accounts ---
        campaign_options_for_multiselect = []
        if not data.empty and 'campaign_name' in data.columns:
            data_for_campaign_options = data # Default to all data
            if selected_accounts and "All Accounts" not in selected_accounts:
                if 'account_name' in data.columns:
                    # Filter data to get campaigns only from selected accounts
                    data_for_campaign_options = data[data['account_name'].isin(selected_accounts)]
                else:
                    # If account_name column doesn't exist for filtering, but accounts are selected,
                    # then no campaigns can be derived from this selection.
                    data_for_campaign_options = pd.DataFrame(columns=data.columns) # Effectively empty

            if not data_for_campaign_options.empty: # Check if after filtering, there's still data
                unique_campaigns_list = data_for_campaign_options['campaign_name'].unique().tolist()
                # Filter out NaN or empty strings from campaign names
                campaign_options_for_multiselect = sorted([
                    str(c) for c in unique_campaigns_list if pd.notna(c) and str(c).strip()
                ])
        # If original data is empty or has no campaign_name column, campaign_options_for_multiselect remains []
        # An st.info message for "No data loaded..." is shown if data is empty.
        
        if data.empty and ('campaign_name' not in data.columns or not campaign_options_for_multiselect):
             st.sidebar.info("No data loaded to determine campaigns for filtering.")


        # --- Campaign Multiselect with "Select All" ---
        # Initialize the session state for selected campaigns if it doesn't exist
        if 'selected_campaigns_key' not in st.session_state:
            st.session_state.selected_campaigns_key = [] # Default to no campaigns selected

        if campaign_options_for_multiselect: # Only show "Select All" if there are campaign options
            # Callback for the "Select All" checkbox
            def on_toggle_select_all_campaigns():
                if st.session_state.get('select_all_campaigns_checkbox_key', False):
                    # If "Select All" is checked, set multiselect state to all options
                    st.session_state.selected_campaigns_key = campaign_options_for_multiselect[:]
                else:
                    # If "Select All" is unchecked, clear multiselect state
                    st.session_state.selected_campaigns_key = []
            
            # Determine the current state of the "Select All" checkbox
            # It should be checked if all available campaigns are currently in st.session_state.selected_campaigns_key
            # and there are campaigns to select.
            all_currently_selected = False
            if campaign_options_for_multiselect: # Ensure there are options to compare against
                if set(st.session_state.get('selected_campaigns_key', [])) == set(campaign_options_for_multiselect):
                    all_currently_selected = True
            
            st.checkbox(
                "Select/Deselect All Campaigns",
                value=all_currently_selected,
                key='select_all_campaigns_checkbox_key',
                on_change=on_toggle_select_all_campaigns,
                help="Toggle to select or deselect all campaigns currently available in the list below."
            )

        # The multiselect widget for campaigns.
        # Its state is directly managed by st.session_state.selected_campaigns_key.
        selected_campaigns = st.multiselect(
            "Campaign(s)",
            options=campaign_options_for_multiselect,
            key='selected_campaigns_key' 
            # The 'default' parameter is implicitly handled by initializing 
            # st.session_state.selected_campaigns_key before this widget.
        )
        
        numeric_columns = data.select_dtypes(include='number').columns.tolist()

        # Define desired default metrics
        desired_default_main_metrics = ["gross_revenue"] # Updated default
        desired_default_side_metrics = ["orders_(sku)", "cost_per_order", "cost", "roi"] # Updated defaults

        # Filter defaults to only include available numeric columns
        actual_default_main_metrics = [m for m in desired_default_main_metrics if m in numeric_columns]
        actual_default_side_metrics = [m for m in desired_default_side_metrics if m in numeric_columns]
        
        main_metrics = st.multiselect(
        "📈 Main Graph & KPI Metrics",
        options=numeric_columns,
        default=actual_default_main_metrics
        )

        side_metrics = st.multiselect(
            "📊 Side-by-Side Metric Charts",
            options=numeric_columns,
            default=actual_default_side_metrics
        )


    # Filter data
    if data.empty: # If no data after loading and potential placeholder creation, show message
        st.info("No data available to display. Please upload reports or check Google Sheet.")
        st.stop()

    base_mask = (
        (data['report_date'] >= pd.to_datetime(date_range[0])) &
        (data['report_date'] <= pd.to_datetime(date_range[1]))
    )

    # Apply campaign filter
    if 'campaign_name' in data.columns: # Ensure campaign_name column exists in the main data
        base_mask &= data['campaign_name'].isin(selected_campaigns)
    # If selected_campaigns is empty (default or no valid options after account filter),
    # .isin([]) correctly results in a False mask for this part, filtering out all campaigns.

    # Apply account filter
    if 'account_name' in data.columns: # Ensure column exists
        if selected_accounts: # If something is selected in the account filter
            if "All Accounts" not in selected_accounts: # If "All Accounts" is not selected, filter by the selected accounts
                base_mask &= data['account_name'].isin(selected_accounts)
            # If "All Accounts" IS selected (it might be the only one, or with others),
            # we effectively don't filter by specific accounts here, showing all.
            # The presence of "All Accounts" overrides specific selections for this logic.
        else: # If selected_accounts is empty (the new default)
            # Filter out all data based on accounts, as the filter is empty.
            base_mask &= data['account_name'].isin([]) # Results in a False mask for this part
    
    filtered_data = data[base_mask]

    # Group for chart
    if not filtered_data.empty and main_metrics:
        grouped = (
            filtered_data[['report_date', 'campaign_name'] + main_metrics]
            .groupby(['report_date', 'campaign_name'])
            .sum()
            .reset_index()
        )

        melted = grouped.melt(
            id_vars=['report_date', 'campaign_name'],
            value_vars=main_metrics,
            var_name='Metric',
            value_name='Value'
        )

        line_chart = alt.Chart(melted).mark_line(point=True).encode(
            x='report_date:T',
            y='Value:Q',
            color=alt.Color('campaign_name:N', legend=alt.Legend(title="Campaign")),
            strokeDash='Metric:N',
            tooltip=['report_date:T', 'campaign_name:N', 'Metric:N', 'Value:Q']
        ).properties(
            width=900,
            height=450,
            title="📈 Campaign Performance Over Time"
        ).interactive()

        st.altair_chart(line_chart, use_container_width=True)


    # Side-by-side metric charts (wider layout for full screen)
    if not filtered_data.empty and side_metrics:
        st.subheader("📊 Side-by-Side Metric Charts by Campaign")

        # Break metrics into chunks of 2 per row for better visual spacing
        metric_chunks = [side_metrics[i:i+2] for i in range(0, len(side_metrics), 2)]

        for chunk in metric_chunks:
            cols = st.columns(len(chunk))  # 2-column layout
            for i, metric in enumerate(chunk):
                with cols[i]:
                    chart_data = (
                        filtered_data[['report_date', 'campaign_name', metric]]
                        .groupby(['report_date', 'campaign_name'])
                        .sum()
                        .reset_index()
                    )

                    chart = alt.Chart(chart_data).mark_line(point=True).encode(
                        x=alt.X('report_date:T', title=''),
                        y=alt.Y(metric, title=metric.replace("_", " ").title(), scale=alt.Scale(zero=True)),
                        color=alt.Color('campaign_name:N', legend=None),
                        tooltip=[
                            alt.Tooltip('report_date:T', title="Date"),
                            alt.Tooltip('campaign_name:N', title="Campaign"),
                            alt.Tooltip(metric, title=metric.replace("_", " ").title(), format=",.2f")
                        ]
                    ).properties(
                        width=600,  # ← wider width
                        height=300,
                        title=metric.replace("_", " ").title()
                    )

                    st.altair_chart(chart, use_container_width=False)

    # KPI Summary
    st.subheader("📌 Summary Metrics")

    if not filtered_data.empty:
        # Ensure all required metric columns exist before trying to sum or mean them
        kpi_metrics = ['cost', 'gross_revenue', 'orders_(sku)', 'roi', 'cost_per_order']
        existing_kpi_metrics = [m for m in kpi_metrics if m in filtered_data.columns]

        total_cost = filtered_data['cost'].sum() if 'cost' in existing_kpi_metrics else 0
        total_revenue = filtered_data['gross_revenue'].sum() if 'gross_revenue' in existing_kpi_metrics else 0
        total_orders = filtered_data['orders_(sku)'].sum() if 'orders_(sku)' in existing_kpi_metrics else 0
        avg_roi = (total_revenue / total_cost) if total_cost > 0 else 0  # Calculate ROI based on total cost and revenue
        avg_cpo = filtered_data['cost_per_order'].mean() if 'cost_per_order' in existing_kpi_metrics else 0
        
        # Handle potential NaN from .mean() if all values were NaN (e.g. for a single row with NaN)
        avg_cpo = 0 if pd.isna(avg_cpo) else avg_cpo

        col1, col2, col3 = st.columns(3)
        col1.metric("💰 Total Cost", f"${total_cost:,.2f}")
        col2.metric("📈 Gross Revenue", f"${total_revenue:,.2f}")
        col3.metric("📊 Avg ROI/MER", f"{avg_roi:.2f}x")

        col4, col5 = st.columns(2)
        col4.metric("📦 Total Orders", f"{int(total_orders):,}")
        col5.metric("🧾 Avg Cost/Order", f"${avg_cpo:,.2f}")

    else:
        st.info("No data to summarize for the current filter selection.")

    # 📤 Export Filtered Data to Excel File
    st.subheader("📤 Export Report")

    if not filtered_data.empty:
        export_granularity = st.selectbox(
            "🗂 Grouping for Export",
            options=["Daily", "Weekly", "Monthly"],
            index=1,
            help="Aggregate data by day, week (Monday start), or calendar month"
        )

        # Ensure datetime
        filtered_data['report_date'] = pd.to_datetime(filtered_data['report_date'])

        # Assign period
        if export_granularity == "Weekly":
            filtered_data['period'] = filtered_data['report_date'].dt.to_period("W").apply(lambda r: r.start_time)
        elif export_granularity == "Monthly":
            filtered_data['period'] = filtered_data['report_date'].dt.to_period("M").apply(lambda r: r.start_time)
        else:
            filtered_data['period'] = filtered_data['report_date'].dt.date

        # Aggregate numeric data by account, campaign + period
        grouping_cols = ['account_name', 'campaign_name', 'period']
        # Ensure all grouping columns exist in filtered_data
        existing_grouping_cols = [col for col in grouping_cols if col in filtered_data.columns]
        
        numeric_columns = filtered_data.select_dtypes(include='number').columns.tolist()
        
        if not all(col in filtered_data.columns for col in existing_grouping_cols):
            st.error("One or more grouping columns are missing from the filtered data. Cannot generate export.")
            summary_df = pd.DataFrame() # Empty df
        else:
            summary_df = (
                filtered_data
                .groupby(existing_grouping_cols)[numeric_columns]
                .sum()
                .reset_index()
            )

            # Recalculate ROI and Cost/Order if necessary columns exist
            if 'gross_revenue' in summary_df.columns and 'cost' in summary_df.columns:
                summary_df['roi'] = summary_df['gross_revenue'] / summary_df['cost']
            else:
                summary_df['roi'] = 0
            
            if 'cost' in summary_df.columns and 'orders_(sku)' in summary_df.columns and summary_df['orders_(sku)'].ne(0).any():
                 summary_df['cost_per_order'] = summary_df['cost'] / summary_df['orders_(sku)']
            else:
                summary_df['cost_per_order'] = 0


            # Replace infinite or undefined values
            summary_df.replace([float('inf'), -float('inf')], pd.NA, inplace=True)
            summary_df['roi'] = summary_df['roi'].fillna(0).round(2)
            summary_df['cost_per_order'] = summary_df['cost_per_order'].fillna(0).round(2)
            summary_df['period'] = pd.to_datetime(summary_df['period']).dt.strftime('%Y-%m-%d')


        # Define currency columns
        currency_columns = ['cost', 'gross_revenue', 'cost_per_order', 'net_cost']

        # Export to Excel (one sheet per campaign)
        output = io.BytesIO()
        if not summary_df.empty:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # --- Create "All Data" sheet FIRST ---
                all_data_sheet_name = "All Data"
                # Use a copy of summary_df for the "All Data" sheet to avoid modifying original for campaign sheets
                all_data_df_export = summary_df.copy() 
                
                # Ensure 'account_name' is in the columns for grouping if it exists
                group_by_cols_all_data = []
                if 'account_name' in all_data_df_export.columns:
                    group_by_cols_all_data.append('account_name')
                if 'campaign_name' in all_data_df_export.columns:
                    group_by_cols_all_data.append('campaign_name')

                if not group_by_cols_all_data: # Should not happen if summary_df was created correctly
                    st.warning("Cannot create 'All Data' sheet as grouping columns are missing.")
                else:
                    # Create the sheet
                    # To write to the sheet, we'll append dataframes directly
                    # For headers and formatting, we'll need to access the worksheet
                    # Write an empty df first to create the sheet, then get the worksheet object
                    pd.DataFrame().to_excel(writer, sheet_name=all_data_sheet_name) # Create empty sheet
                    worksheet_all_data = writer.sheets[all_data_sheet_name]
                    
                    current_row_all_data = 1 # openpyxl is 1-indexed
                    dandelion_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") # Gold/Dandelion

                    # Sort summary_df by account_name, then campaign_name for ordered output
                    if 'account_name' in all_data_df_export.columns and 'campaign_name' in all_data_df_export.columns:
                        all_data_df_export_sorted = all_data_df_export.sort_values(by=['account_name', 'campaign_name', 'period'])
                    elif 'campaign_name' in all_data_df_export.columns:
                         all_data_df_export_sorted = all_data_df_export.sort_values(by=['campaign_name', 'period'])
                    else:
                        all_data_df_export_sorted = all_data_df_export # No specific sort if key columns missing

                    # Write data for each campaign block
                    for group_keys, campaign_group_data in all_data_df_export_sorted.groupby(group_by_cols_all_data, sort=False):
                        # Write Header for this campaign block
                        header_list = campaign_group_data.columns.tolist()
                        for col_idx, header_name in enumerate(header_list, 1):
                            cell = worksheet_all_data.cell(row=current_row_all_data, column=col_idx, value=header_name)
                            cell.fill = dandelion_fill
                        
                        #worksheet_all_data.freeze_panes = worksheet_all_data.cell(row=current_row_all_data + 1, column=1)
                        current_row_all_data += 1

                        # Write Data for this campaign block
                        for r_idx, record in enumerate(campaign_group_data.to_dict(orient='records')):
                            for c_idx, col_name in enumerate(header_list, 1):
                                worksheet_all_data.cell(row=current_row_all_data + r_idx, column=c_idx, value=record[col_name])
                        
                        # Apply formatting for the current block of data
                        for col_idx_format, column_name_format in enumerate(header_list, 1):
                            col_letter_format = get_column_letter(col_idx_format)
                            max_len = max(
                                campaign_group_data[column_name_format].astype(str).map(len).max(skipna=True),
                                len(column_name_format)
                            ) + 2
                            if column_name_format == "period": max_len = 20
                            worksheet_all_data.column_dimensions[col_letter_format].width = max(worksheet_all_data.column_dimensions[col_letter_format].width or 0, max_len)

                            if column_name_format in currency_columns:
                                for row_idx_format in range(current_row_all_data, current_row_all_data + len(campaign_group_data)):
                                    cell = worksheet_all_data.cell(row=row_idx_format, column=col_idx_format)
                                    cell.number_format = '$#,##0.00'
                        
                        current_row_all_data += len(campaign_group_data)
                        current_row_all_data += 1 # Add a blank row as a separator

                # --- Create individual campaign sheets (existing logic) ---
                # Ensure 'campaign_name' exists for this grouping
                if 'campaign_name' in summary_df.columns:
                    for campaign_name_key, campaign_data in summary_df.groupby('campaign_name'):
                        sheet_name = str(campaign_name_key)[:31] # Ensure sheet name is string
                        # Select relevant columns for individual sheets (e.g., drop account_name if preferred, or keep it)
                        # For now, keeping all columns from summary_df
                        campaign_data_to_export = campaign_data.copy()
                        campaign_data_to_export.to_excel(writer, index=False, sheet_name=sheet_name)

                        # Get worksheet and auto-adjust column widths
                        workbook = writer.book
                        worksheet = writer.sheets[sheet_name]

                        # ✅ Freeze header row
                        worksheet.freeze_panes = worksheet["A2"]

                        # Highlight headers with light green
                        header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        for cell in worksheet[1]:  # Row 1
                            cell.fill = header_fill

                        # Apply currency formatting and auto-adjust column widths
                        for col_idx, column_name in enumerate(campaign_data_to_export.columns, 1):
                            col_letter = get_column_letter(col_idx)
                            
                            # Auto-adjust column widths
                            if column_name == "period":
                                max_length = 20  # wider default for date strings
                            else:
                                max_length = max(
                                    campaign_data_to_export[column_name].astype(str).map(len).max(skipna=True), # Added skipna=True
                                    len(str(column_name)) # Ensure column_name is string for len()
                                ) + 2
                            worksheet.column_dimensions[col_letter].width = max_length

                            # Apply currency format
                            if column_name in currency_columns:
                                for row_idx in range(2, worksheet.max_row + 1): # Start from row 2 (below header)
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    cell.number_format = '$#,##0.00'
                else:
                    st.warning("Skipping individual campaign sheets as 'campaign_name' column is missing in summary.")

        else: # summary_df is empty
            st.info("No summary data to export.")
            # Still need to provide a download button for an empty file or handle this case
            # For now, if summary_df is empty, the output BytesIO will be empty,
            # and download button might download an empty/corrupt file.
            # A better approach would be to not show download button or show a message.
            # However, the outer `if not filtered_data.empty:` handles this.

        # Download button
        if not summary_df.empty: # Only show download if there's data
            start_str = date_range[0].strftime("%Y-%m-%d")
            end_str = date_range[1].strftime("%Y-%m-%d")
            file_name = f"tiktok_summary_{export_granularity.lower()}_{start_str}_to_{end_str}.xlsx"

            st.download_button(
                label="📥 Download Aggregated Excel Report",
                data=output.getvalue(),
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("No data to export for the selected filters.")

    # 📧 Send Weekly Email Data (GCS Upload & Zapier) - Admin Only
    if st.session_state.get("is_admin"):
        st.subheader("📧 Send Weekly Email Data")

        # Condition to show the button: filtered_data, summary_df, start_str, end_str, and output (Excel bytes) must be available
        if not filtered_data.empty and 'summary_df' in locals() and 'start_str' in locals() and 'end_str' in locals() and 'output' in locals() and 'file_name' in locals():
            if st.button("🚀 Send Data to Zapier for Weekly Email"):
                if "gcp" not in st.secrets or "gcs_bucket_name" not in st.secrets:
                    st.error("GCP credentials or GCS bucket name not found in secrets.")
                elif "zapier_webhook_url" not in st.secrets:
                    st.error("Zapier webhook URL not found in secrets.")
                else:
                    with st.spinner("Preparing charts and sending data to Zapier..."):
                        # Prepare data for the 5 specific charts
                        # Ensure 'orders_(sku)' is the correct column name from your data
                        required_metrics_for_charts = {
                            'cost': 'Cost',
                            'gross_revenue': 'Gross Revenue',
                            'orders_(sku)': 'Orders', # Assuming 'orders' in prompt means 'orders_(sku)'
                        }
                        
                        # Aggregate data for sum-based metrics
                        zapier_chart_data_agg = filtered_data.groupby(['report_date', 'campaign_name']).agg(
                            **{metric: (metric, 'sum') for metric in required_metrics_for_charts.keys()}
                        ).reset_index()

                        # Calculate ROI and CPO for these aggregated groups
                        zapier_chart_data_agg['roi'] = (zapier_chart_data_agg['gross_revenue'] / zapier_chart_data_agg['cost'])
                        zapier_chart_data_agg['cost_per_order'] = (zapier_chart_data_agg['cost'] / zapier_chart_data_agg['orders_(sku)'])
                        
                        # Handle potential NaN/inf values from division
                        zapier_chart_data_agg.replace([float('inf'), -float('inf')], pd.NA, inplace=True) # Use pd.NA for consistency
                        zapier_chart_data_agg['roi'] = zapier_chart_data_agg['roi'].fillna(0)
                        zapier_chart_data_agg['cost_per_order'] = zapier_chart_data_agg['cost_per_order'].fillna(0)

                        target_charts_to_upload = {
                            "cost": "Cost",
                            "gross_revenue": "Gross Revenue",
                            "roi": "ROI",
                            "orders_(sku)": "Orders", # Maps to 'orders_(sku)' column
                            "cost_per_order": "Cost Per Order"
                        }
                        
                        zapier_chart_urls = {}
                        all_charts_uploaded = True
                        excel_report_url = None

                        # 1. Upload Excel Report
                        st.write("Uploading Excel report to GCS...")
                        excel_bytes_to_upload = output.getvalue() # Get bytes from the existing 'output' BytesIO
                        excel_report_url = upload_excel_to_gcs(
                            excel_bytes_to_upload,
                            st.secrets["gcs_bucket_name"],
                            st.secrets["gcp"],
                            original_filename=file_name # Use the generated file_name
                        )

                        if not excel_report_url:
                            st.error("Failed to upload Excel report. Aborting Zapier send.")
                            all_charts_uploaded = False # Use this flag to prevent further processing
                        else:
                            st.write(f"Excel report uploaded: {excel_report_url}")

                        # 2. Upload Charts (only if Excel upload was successful)
                        if all_charts_uploaded: # Check if Excel upload was successful
                            for metric_col, display_name in target_charts_to_upload.items():
                                st.write(f"Generating chart for {display_name}...")
                                chart_obj = generate_specific_metric_chart(
                                    chart_df=zapier_chart_data_agg, # Use the aggregated and calculated data
                                    metric_col_name=metric_col,
                                    display_metric_name=display_name
                                )
                                if chart_obj:
                                    st.write(f"Uploading {display_name} chart to GCS...")
                                    image_url = upload_chart_to_gcs(
                                        chart_obj,
                                        st.secrets["gcs_bucket_name"],
                                        st.secrets["gcp"]
                                    )
                                    if image_url:
                                        zapier_chart_urls[f"{metric_col.replace('_(sku)', '')}_url"] = image_url # e.g., orders_url
                                        st.write(f"{display_name} chart uploaded: {image_url}")
                                    else:
                                        st.error(f"Failed to upload {display_name} chart.")
                                        all_charts_uploaded = False
                                        break # Stop if one chart fails
                                else:
                                    st.error(f"Failed to generate chart for {display_name}.")
                                    all_charts_uploaded = False
                                    break # Stop if one chart fails to generate


                        if all_charts_uploaded and excel_report_url and len(zapier_chart_urls) == len(target_charts_to_upload):
                            # Prepare payload for Zapier
                            payload = {
                                "start_date": start_str, # Defined in export section
                                "end_date": end_str,     # Defined in export section
                                "summary_data": summary_df.to_dict(orient="records"), # summary_df from export section
                                "excel_report_url": excel_report_url, # Add Excel report URL
                                "chart_images": zapier_chart_urls
                            }

                            # Send to Zapier
                            try:
                                response = requests.post(st.secrets["zapier_webhook_url"], json=payload)
                                response.raise_for_status() # Raise an exception for HTTP errors
                                st.success("✅ Data and chart images successfully sent to Zapier!")
                                st.balloons()
                            except requests.exceptions.RequestException as e:
                                st.error(f"❌ Failed to send data to Zapier: {e}")
                                if 'response' in locals() and response is not None:
                                    st.error(f"Zapier response: {response.text}")
                                else:
                                    st.error("No response object from Zapier.")
                        elif not excel_report_url:
                            # Error already shown by upload_excel_to_gcs or the check above
                            pass 
                        elif not all_charts_uploaded:
                            st.error("❌ Not all charts were uploaded successfully. Data not sent to Zapier.")
                        else:
                            st.error("❌ Chart generation or upload failed for some charts. Data not sent to Zapier.")
        elif 'summary_df' not in locals() or 'output' not in locals():
            st.info("Generate an export report first (which creates summary data and the Excel file) to enable sending data to Zapier.")
        else: # filtered_data is empty
            st.info("No data available for the current filters to send to Zapier.")

    # Historical data query
    if st.button('📊 View Historical Data'):
        records = sh.get_all_records()
        history_df = pd.DataFrame(records)
        st.write("Historical Ad Report Data:")
        st.dataframe(history_df)

# --- Authentication Flow ---
ENABLE_AUTH = os.getenv("STREAMLIT_ENABLE_AUTH", "false").lower() == "true"

if not ENABLE_AUTH:
    # Authentication is disabled, grant full access by default for local development
    if "authenticated" not in st.session_state: # Initialize if not already set by a previous run
        st.session_state.authenticated = True
        st.session_state.is_admin = True
    st.sidebar.info("🔑 Authentication Disabled (Dev Mode)")
    run_main_app()
else:
    # Authentication is enabled
    if "authenticated" not in st.session_state:
        st.title("🔒 Protected Dashboard")
        with st.form("auth_form"):
            pw = st.text_input("Enter viewer password", type="password")
            admin_key = st.text_input("Enter admin key (optional)", type="password")
            submitted = st.form_submit_button("Login")

        if submitted:
            if pw == st.secrets["app_password"]:
                st.session_state.authenticated = True
                st.session_state.is_admin = admin_key == st.secrets["admin_key"]
                st.rerun()
            else:
                st.error("Incorrect password")
        st.stop() # Stop execution if login form is shown and not successfully submitted
    
    # If we reach here, authentication is enabled AND user is authenticated.
    
    # Sidebar elements for authenticated users when auth is enabled
    with st.sidebar:
        if st.button("🔒 Logout"):
            if "authenticated" in st.session_state:
                del st.session_state.authenticated
            if "is_admin" in st.session_state:
                del st.session_state.is_admin
            st.rerun()

        # Allow non-admins to enter admin key later (only if auth is enabled)
        # Ensure is_admin exists before checking its value
        if "is_admin" not in st.session_state: # Should be set by login, but as a safeguard
            st.session_state.is_admin = False
            
        if st.session_state.authenticated and not st.session_state.is_admin:
            with st.expander("🔑 Admin Access"):
                with st.form("admin_key_form"):
                    admin_key_input = st.text_input("Enter admin key to enable uploads", type="password")
                    admin_key_submitted = st.form_submit_button("Unlock Admin Features")

                if admin_key_submitted:
                    if admin_key_input == st.secrets["admin_key"]:
                        st.session_state.is_admin = True
                        st.sidebar.success("Admin features unlocked!")
                        st.rerun()
                    elif admin_key_input: # if a key was entered but it's wrong
                        st.sidebar.error("Incorrect admin key.")
    
    # Ensure default state for admin if not explicitly set after login/admin key attempt
    if "is_admin" not in st.session_state:
        st.session_state.is_admin = False

    run_main_app()
