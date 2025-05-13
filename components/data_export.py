import streamlit as st
import pandas as pd
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def render_export_section(filtered_data):
    """Render export section and return the export data for possible further use."""
    st.subheader("📤 Export Report")
    
    if not filtered_data.empty:
        export_granularity = st.selectbox(
            "🗂 Grouping for Export",
            options=["Daily", "Weekly", "Monthly"],
            index=1,
            help="Aggregate data by day, week (Monday start), or calendar month"
        )

        # Ensure datetime
        filtered_data['report_date'] = pd.to_datetime(filtered_data['report_date'])

        # Assign period
        if export_granularity == "Weekly":
            filtered_data['period'] = filtered_data['report_date'].dt.to_period("W").apply(lambda r: r.start_time)
        elif export_granularity == "Monthly":
            filtered_data['period'] = filtered_data['report_date'].dt.to_period("M").apply(lambda r: r.start_time)
        else:
            filtered_data['period'] = filtered_data['report_date'].dt.date

        # Process export data and generate Excel
        summary_df, output, file_name = process_export_data(filtered_data, export_granularity)
        
        # Download button
        if not summary_df.empty:  # Only show download if there's data
            st.download_button(
                label="📥 Download Aggregated Excel Report",
                data=output.getvalue(),
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Return export data for potential further use (like email)
            return {
                "summary_df": summary_df,
                "output": output,
                "file_name": file_name,
                "export_granularity": export_granularity
            }
    else:
        st.info("No data to export for the selected filters.")
    
    return None

def process_export_data(filtered_data, export_granularity):
    """Process data for export and create Excel file."""
    # Aggregate numeric data by account, campaign + period
    grouping_cols = ['account_name', 'campaign_name', 'period']
    existing_grouping_cols = [col for col in grouping_cols if col in filtered_data.columns]
    
    numeric_columns = filtered_data.select_dtypes(include='number').columns.tolist()
    
    if not all(col in filtered_data.columns for col in existing_grouping_cols):
        st.error("One or more grouping columns are missing from the filtered data. Cannot generate export.")
        return pd.DataFrame(), io.BytesIO(), ""
    
    summary_df = (
        filtered_data
        .groupby(existing_grouping_cols)[numeric_columns]
        .sum()
        .reset_index()
    )

    # Recalculate ROI and Cost/Order
    if 'gross_revenue' in summary_df.columns and 'cost' in summary_df.columns:
        summary_df['roi'] = summary_df['gross_revenue'] / summary_df['cost']
    else:
        summary_df['roi'] = 0
    
    if 'cost' in summary_df.columns and 'orders_(sku)' in summary_df.columns and summary_df['orders_(sku)'].ne(0).any():
         summary_df['cost_per_order'] = summary_df['cost'] / summary_df['orders_(sku)']
    else:
        summary_df['cost_per_order'] = 0

    # Clean up values
    summary_df.replace([float('inf'), -float('inf')], pd.NA, inplace=True)
    summary_df['roi'] = summary_df['roi'].fillna(0).round(2)
    summary_df['cost_per_order'] = summary_df['cost_per_order'].fillna(0).round(2)
    summary_df['period'] = pd.to_datetime(summary_df['period']).dt.strftime('%Y-%m-%d')

    # Define currency columns
    currency_columns = ['cost', 'gross_revenue', 'cost_per_order', 'net_cost']

    # Generate Excel file
    output = io.BytesIO()
    
    if not summary_df.empty:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            create_all_data_sheet(writer, summary_df, currency_columns)
            create_campaign_sheets(writer, summary_df, currency_columns)
    
    # Generate filename
    start_date = filtered_data['report_date'].min().strftime("%Y-%m-%d")
    end_date = filtered_data['report_date'].max().strftime("%Y-%m-%d")
    file_name = f"tiktok_summary_{export_granularity.lower()}_{start_date}_to_{end_date}.xlsx"
    
    return summary_df, output, file_name

def create_all_data_sheet(writer, summary_df, currency_columns):
    """Create the 'All Data' sheet in the Excel workbook."""
    all_data_sheet_name = "All Data"
    all_data_df_export = summary_df.copy()
    
    group_by_cols_all_data = []
    if 'account_name' in all_data_df_export.columns:
        group_by_cols_all_data.append('account_name')
    if 'campaign_name' in all_data_df_export.columns:
        group_by_cols_all_data.append('campaign_name')

    if not group_by_cols_all_data:
        st.warning("Cannot create 'All Data' sheet as grouping columns are missing.")
        return
    
    pd.DataFrame().to_excel(writer, sheet_name=all_data_sheet_name)
    worksheet_all_data = writer.sheets[all_data_sheet_name]
    
    current_row_all_data = 1
    dandelion_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Sort data
    if 'account_name' in all_data_df_export.columns and 'campaign_name' in all_data_df_export.columns:
        all_data_df_export_sorted = all_data_df_export.sort_values(by=['account_name', 'campaign_name', 'period'])
    elif 'campaign_name' in all_data_df_export.columns:
        all_data_df_export_sorted = all_data_df_export.sort_values(by=['campaign_name', 'period'])
    else:
        all_data_df_export_sorted = all_data_df_export

    # Write data by groups
    for group_keys, campaign_group_data in all_data_df_export_sorted.groupby(group_by_cols_all_data, sort=False):
        # Write Header
        header_list = campaign_group_data.columns.tolist()
        for col_idx, header_name in enumerate(header_list, 1):
            cell = worksheet_all_data.cell(row=current_row_all_data, column=col_idx, value=header_name)
            cell.fill = dandelion_fill
        
        current_row_all_data += 1

        # Write Data
        for r_idx, record in enumerate(campaign_group_data.to_dict(orient='records')):
            for c_idx, col_name in enumerate(header_list, 1):
                worksheet_all_data.cell(row=current_row_all_data + r_idx, column=c_idx, value=record[col_name])
        
        # Apply formatting
        for col_idx_format, column_name_format in enumerate(header_list, 1):
            col_letter_format = get_column_letter(col_idx_format)
            max_len = max(
                campaign_group_data[column_name_format].astype(str).map(len).max(skipna=True),
                len(column_name_format)
            ) + 2
            if column_name_format == "period": max_len = 20
            worksheet_all_data.column_dimensions[col_letter_format].width = max(
                worksheet_all_data.column_dimensions[col_letter_format].width or 0, max_len
            )

            if column_name_format in currency_columns:
                for row_idx_format in range(current_row_all_data, current_row_all_data + len(campaign_group_data)):
                    cell = worksheet_all_data.cell(row=row_idx_format, column=col_idx_format)
                    cell.number_format = '$#,##0.00'
        
        current_row_all_data += len(campaign_group_data)
        current_row_all_data += 1  # Add a blank row as a separator

def create_campaign_sheets(writer, summary_df, currency_columns):
    """Create individual sheets for each campaign."""
    if 'campaign_name' not in summary_df.columns:
        st.warning("Skipping individual campaign sheets as 'campaign_name' column is missing in summary.")
        return
    
    for campaign_name_key, campaign_data in summary_df.groupby('campaign_name'):
        sheet_name = str(campaign_name_key)[:31]  # Excel has a 31-character sheet name limit
        campaign_data_to_export = campaign_data.copy()
        campaign_data_to_export.to_excel(writer, index=False, sheet_name=sheet_name)

        # Get worksheet and adjust formatting
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Freeze header row
        worksheet.freeze_panes = worksheet["A2"]

        # Highlight headers with light green
        header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        for cell in worksheet[1]:  # Row 1
            cell.fill = header_fill

        # Apply currency formatting and auto-adjust column widths
        for col_idx, column_name in enumerate(campaign_data_to_export.columns, 1):
            col_letter = get_column_letter(col_idx)
            
            # Auto-adjust column widths
            if column_name == "period":
                max_length = 20  # wider default for date strings
            else:
                max_length = max(
                    campaign_data_to_export[column_name].astype(str).map(len).max(skipna=True),
                    len(str(column_name))
                ) + 2
            worksheet.column_dimensions[col_letter].width = max_length

            # Apply currency format
            if column_name in currency_columns:
                for row_idx in range(2, worksheet.max_row + 1):  # Start from row 2 (below header)
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = '$#,##0.00'